"""Builds the professional Excel scoring + financial model workbook at
outputs/scoring_model.xlsx. Six tabs: City Scoring, Scoring Thresholds,
Sensitivity Analysis, one financial-model tab per recommended city, and
Data Sources.
"""

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

RAW_DIR = PROJECT_ROOT / "data_collection" / "raw"
OUTPUTS_DIR = PROJECT_ROOT / "outputs"

HEADER_FILL = PatternFill("solid", fgColor="1565C0")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="0B0B0B")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="52514E")
RECOMMENDED_FILL = PatternFill("solid", fgColor="C8E6C9")
NOT_RECOMMENDED_FILL = PatternFill("solid", fgColor="F5F5F5")
AMBER_FILL = PatternFill("solid", fgColor="FFE0B2")
GREEN_FILL = PatternFill("solid", fgColor="C8E6C9")
THIN_BORDER = Side(style="thin", color="C3C2B7")
CELL_BORDER = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)
THICK_BOTTOM = Border(bottom=Side(style="medium", color="0B0B0B"))


def _style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER


def _autofit_columns(ws, widths: dict[str, int]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def build_city_scoring_tab(wb: Workbook, scores: pd.DataFrame):
    ws = wb.create_sheet("City Scoring")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Meridian Home — Market Entry Scoring Model"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Prepared by Balraj Kooner | Masters in Business Analytics"
    ws["A2"].font = SUBTITLE_FONT

    headers = [
        "City", "Rank",
        "Dim1: Pop & Growth", "Dim2: Income & Spend", "Dim3: Homeownership & Housing",
        "Dim4: Competitive Saturation", "Dim5: Retail Lease Cost", "Dim6: Labor Cost",
        "Dim7: Logistics Proximity", "Composite Score", "Recommendation",
    ]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    scores_sorted = scores.sort_values("rank")
    for i, (_, row) in enumerate(scores_sorted.iterrows()):
        r = header_row + 1 + i
        values = [
            row["metro_name"], int(row["rank"]),
            row["dim1_score"], row["dim2_score"], row["dim3_score"], row["dim4_score"],
            row["dim5_score"], row["dim6_score"], row["dim7_score"], row["composite_score"],
            "RECOMMENDED" if row["recommended"] else "",
        ]
        fill = RECOMMENDED_FILL if row["recommended"] else NOT_RECOMMENDED_FILL
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill = fill
            cell.border = CELL_BORDER
            if col >= 3:
                cell.alignment = Alignment(horizontal="right")
                if col in (3, 4, 5, 6, 7, 8, 9, 10):
                    cell.number_format = "0.00"

    ws.freeze_panes = "A5"
    _autofit_columns(
        ws,
        {
            "A": 18, "B": 6, "C": 16, "D": 16, "E": 20, "F": 20,
            "G": 18, "H": 14, "I": 16, "J": 14, "K": 16,
        },
    )


def build_scoring_thresholds_tab(wb: Workbook):
    ws = wb.create_sheet("Scoring Thresholds")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Scoring Thresholds — How Raw Data Becomes a 1-5 Score"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Thresholds set against the real range of values across the 5 candidate metros"
    ws["A2"].font = SUBTITLE_FONT

    sections = [
        ("Dimension 1 — Population & Growth (weight 15%)", [
            ("Population: >2.5M", "5"), ("1.5M-2.5M", "4"), ("1.0M-1.5M", "3"),
            ("0.75M-1.0M", "2"), ("<0.75M", "1"),
            ("Growth rate: >15%", "5"), ("10-15%", "4"), ("7-10%", "3"), ("4-7%", "2"), ("<4%", "1"),
        ]),
        ("Dimension 2 — Income & Furnishing Spend (weight 20%)", [
            ("Income: >$90K", "5"), ("$75-90K", "4"), ("$65-75K", "3"), ("$55-65K", "2"), ("<$55K", "1"),
            ("HH furnishing spend: >$2,200", "5"), ("$1,800-2,200", "4"),
            ("$1,500-1,800", "3"), ("$1,200-1,500", "2"), ("<$1,200", "1"),
        ]),
        ("Dimension 3 — Homeownership & Housing Activity (weight 20%)", [
            ("Homeownership: >70%", "5"), ("65-70%", "4"), ("60-65%", "3"), ("55-60%", "2"), ("<55%", "1"),
            ("5yr appreciation: >25%", "5"), ("15-25%", "4"),
            ("0-15%", "3"), ("-10-0%", "2"), ("<-10%", "1"),
            ("Building permits: >20,000", "5"), ("15,000-20,000", "4"),
            ("10,000-15,000", "3"), ("7,000-10,000", "2"), ("<7,000", "1"),
        ]),
        ("Dimension 4 — Competitive Saturation (weight 20%, INVERTED)", [
            ("Competitors per 100k: <0.25", "5"), ("0.25-0.30", "4"),
            ("0.30-0.35", "3"), ("0.35-0.40", "2"), (">0.40", "1"),
        ]),
        ("Dimension 5 — Retail Lease Rate (weight 10%, INVERTED)", [
            ("$/sqft/yr: <$23", "5"), ("$23-26", "4"), ("$26-28.50", "3"),
            ("$28.50-31", "2"), (">$31", "1"),
        ]),
        ("Dimension 6 — Labor Cost (weight 10%, INVERTED)", [
            ("Hourly wage: <$16", "5"), ("$16-18", "4"), ("$18-20", "3"), ("$20-23", "2"), (">$23", "1"),
        ]),
        ("Dimension 7 — Logistics Proximity (weight 5%, INVERTED)", [
            ("Distance to DC: <100mi", "5"), ("100-400mi", "4"), ("400-700mi", "3"),
            ("700-1,000mi", "2"), (">1,000mi", "1"),
        ]),
    ]

    row = 4
    for title, rows in sections:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11, color="1565C0")
        row += 1
        ws.cell(row=row, column=1, value="Raw Value Range")
        ws.cell(row=row, column=2, value="Score")
        _style_header_row(ws, row, 2)
        row += 1
        for label, score in rows:
            ws.cell(row=row, column=1, value=label).border = CELL_BORDER
            score_cell = ws.cell(row=row, column=2, value=score)
            score_cell.border = CELL_BORDER
            score_cell.alignment = Alignment(horizontal="center")
            row += 1
        row += 1

    _autofit_columns(ws, {"A": 48, "B": 10})


def build_sensitivity_tab(wb: Workbook, sensitivity: pd.DataFrame):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Sensitivity Analysis — City Rank by Weight Profile"
    ws["A1"].font = TITLE_FONT

    profile_notes = [
        "Baseline: Pop&Growth 15%, Income&Spend 20%, Housing 20%, Competition 20%, Lease 10%, Labor 10%, Logistics 5%",
        "Financial focus: Income&Spend 25% (+5), Lease 15% (+5), Pop&Growth 10% (-5), Competition 15% (-5)",
        "Growth focus: Pop&Growth 25% (+10), Housing 25% (+5), Income&Spend 15% (-5), Competition 15% (-5), Lease 5% (-5)",
    ]
    for i, note in enumerate(profile_notes):
        ws.cell(row=2 + i, column=1, value=note).font = SUBTITLE_FONT

    header_row = 6
    cols = ["Metro"] + [c.replace("_", " ").title() for c in sensitivity.columns if c != "stability"]
    for col, h in enumerate(cols, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(cols))

    sensitivity_sorted = sensitivity.sort_values("baseline")
    for i, (metro, row_data) in enumerate(sensitivity_sorted.iterrows()):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=metro).border = CELL_BORDER
        for j, col in enumerate([c for c in sensitivity.columns if c != "stability"], start=2):
            cell = ws.cell(row=r, column=j, value=int(row_data[col]))
            cell.alignment = Alignment(horizontal="center")
            cell.border = CELL_BORDER

    is_robust = sensitivity["stability"].iloc[0] == "Robust"
    summary_row = header_row + len(sensitivity_sorted) + 2
    ws.cell(row=summary_row, column=1, value="Recommendation Stable:").font = Font(bold=True)
    stable_cell = ws.cell(row=summary_row, column=2, value="YES" if is_robust else "NO")
    stable_cell.font = Font(bold=True, color="006300" if is_robust else "D03B3B")

    _autofit_columns(ws, {"A": 18, "B": 12, "C": 16, "D": 14})


def build_financial_tab(wb: Workbook, city_name: str, financial_data: dict):
    sheet_name = f"{city_name.split(',')[0]} Financial Model"[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{city_name} — 3-Year Store Cluster Financial Model (2 stores)"
    ws["A1"].font = TITLE_FONT

    pl = pd.DataFrame(financial_data["cluster_pl"])
    header_row = 3
    ws.cell(row=header_row, column=1, value="Line Item")
    for i, year in enumerate(pl["year"], start=2):
        ws.cell(row=header_row, column=i, value=f"Year {int(year)}")
    _style_header_row(ws, header_row, len(pl) + 1)

    line_items = [
        ("Revenue", "revenue"), ("COGS", "cogs"), ("Gross Profit", "gross_profit"),
        ("Rent", "rent"), ("Labor", "labor"), ("Marketing", "marketing"), ("G&A", "ga"),
        ("Total Opex", "total_opex"), ("EBITDA", "ebitda"), ("EBITDA Margin %", "ebitda_margin_pct"),
    ]
    row = header_row + 1
    for label, key in line_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=(key == "ebitda"))
        for i, (_, yr_row) in enumerate(pl.iterrows(), start=2):
            cell = ws.cell(row=row, column=i)
            if key == "ebitda_margin_pct":
                cell.value = yr_row[key] / 100
                cell.number_format = "0.0%"
            else:
                cell.value = yr_row[key]
                cell.number_format = "$#,##0"
            if key == "ebitda":
                cell.fill = GREEN_FILL
                cell.font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Pre-Opening Cost (2 stores)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=450_000 * 2).number_format = "$#,##0"
    ws.cell(row=row, column=1).fill = AMBER_FILL
    ws.cell(row=row, column=2).fill = AMBER_FILL
    row += 2

    ws.cell(row=row, column=1, value="Payback Period (months)").font = Font(bold=True)
    payback = financial_data["payback_months"]
    ws.cell(row=row, column=2, value=payback if payback != float("inf") else "Beyond 3yr horizon")
    row += 1
    ws.cell(row=row, column=1, value="NPV @ 10% (base)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=financial_data["npv_base"]).number_format = "$#,##0"
    row += 1
    ws.cell(row=row, column=1, value="NPV @ 12% (risk-adjusted)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=financial_data["npv_risk_adjusted"]).number_format = "$#,##0"
    row += 2

    ws.cell(row=row, column=1, value="Scenario Analysis (3-Year NPV @ 10%)").font = Font(bold=True, size=11, color="1565C0")
    row += 1
    ws.cell(row=row, column=1, value="Pessimistic (revenue -20%, lease +15%)")
    ws.cell(row=row, column=2, value=financial_data["scenario_pessimistic_npv"]).number_format = "$#,##0"
    row += 1
    ws.cell(row=row, column=1, value="Base Case")
    ws.cell(row=row, column=2, value=financial_data["npv_base"]).number_format = "$#,##0"
    row += 1
    ws.cell(row=row, column=1, value="Optimistic (revenue +15%)")
    ws.cell(row=row, column=2, value=financial_data["scenario_optimistic_npv"]).number_format = "$#,##0"

    _autofit_columns(ws, {"A": 34, "B": 16, "C": 16, "D": 16})


def build_data_sources_tab(wb: Workbook):
    ws = wb.create_sheet("Data Sources")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Data Sources and Assumptions Log"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    headers = ["Dimension", "Data Point", "Source", "URL / Note", "Year"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _style_header_row(ws, header_row, len(headers))

    rows = [
        ("Dim 1", "Population, 5yr growth", "US Census Bureau ACS 5-Year", "api.census.gov/data/2023/acs/acs5 (requires free API key)", "2023"),
        ("Dim 2", "Median household income", "US Census Bureau ACS 5-Year", "api.census.gov/data/2023/acs/acs5", "2023"),
        ("Dim 2", "Home furnishings spend/HH", "BLS Consumer Expenditure Survey, Table 1800 (Region of residence)", "bls.gov/cex/tables/geographic/mean.htm", "2023-2024"),
        ("Dim 3", "Homeownership rate", "US Census Bureau ACS 5-Year", "api.census.gov/data/2023/acs/acs5", "2023"),
        ("Dim 3", "Home value appreciation", "Zillow Research, ZHVI", "files.zillowstatic.com/research/public_csvs/zhvi/", "2021-2026"),
        ("Dim 3", "Building permits", "NAHB (Census monthly permit survey, aggregated to metro)", "nahb.org/.../building-permits-by-state-and-metro-area", "FY2025"),
        ("Dim 4", "Competitor store counts", "Brand official store locators (Crate & Barrel, West Elm, Pottery Barn, RH, Williams-Sonoma)", "Counted within each city's Census-defined MSA, July 2026", "2026"),
        ("Dim 5", "Retail lease rate $/sqft", "CommercialCafe (CoStar-backed live listings)", "commercialcafe.com", "2026"),
        ("Dim 6", "Retail labor wage", "BLS Occupational Employment and Wage Statistics (OEWS), SOC 41-2031", "api.bls.gov/publicAPI/v2/timeseries/data/", "2025"),
        ("Dim 7", "Driving distance to DC", "Web-verified driving distances", "Nashville->Charlotte 409mi, Austin->SLC 1291mi, Denver->SLC 521mi, Indianapolis->Charlotte 576mi", "2026"),
        ("Financial", "Revenue/cost model assumptions", "Project team assumptions, benchmarked against specialty-retail staffing and conversion norms", "See README.md", "2026"),
    ]
    for i, (dim, point, source, url, year) in enumerate(rows):
        r = header_row + 1 + i
        for col, val in enumerate([dim, point, source, url, year], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autofit_columns(ws, {"A": 10, "B": 26, "C": 36, "D": 46, "E": 10})


def build_excel_model():
    scores = pd.read_csv(RAW_DIR / "city_scores.csv")
    sensitivity = pd.read_csv(RAW_DIR / "sensitivity_results.csv", index_col=0)
    financial_results = json.loads((RAW_DIR / "financial_results.json").read_text())

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    build_city_scoring_tab(wb, scores)
    build_scoring_thresholds_tab(wb)
    build_sensitivity_tab(wb, sensitivity)

    recommended = scores[scores["recommended"]].sort_values("rank")
    for _, row in recommended.iterrows():
        build_financial_tab(wb, row["metro_name"], financial_results[row["metro_name"]])

    build_data_sources_tab(wb)

    out_path = OUTPUTS_DIR / "scoring_model.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}")
    print(f"Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    build_excel_model()
